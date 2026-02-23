"""
reporter.py — Excel Report Generator.

Produces a multi-sheet Excel workbook designed for operations and finance
teams. The workbook is formatted to match enterprise reporting standards —
colour-coded severity rows, frozen headers, auto-fitted columns, and a
cover sheet with KPI summary boxes.

Sheets:
    1. Summary          — KPI tiles, leakage by rule, top 5 suppliers
    2. Flagged Items    — Full flagged transaction detail with conditional formatting
    3. Statistics       — Rule performance metrics and daily trend table
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import yaml

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Colour palette (consistent with brand / Power BI standard colours)
# ---------------------------------------------------------------------------
COLOURS = {
    "navy":         "1F4E79",
    "dark_red":     "C00000",
    "dark_green":   "375623",
    "gold":         "BF8F00",
    "light_grey":   "F2F2F2",
    "mid_grey":     "D9D9D9",
    "white":        "FFFFFF",
    "critical_row": "FFCCCC",
    "high_row":     "FFE5CC",
    "medium_row":   "FFFFE0",
    "low_row":      "E2EFDA",
    "header_font":  "FFFFFF",
}

SEVERITY_ROW_COLOURS = {
    "Critical": COLOURS["critical_row"],
    "High":     COLOURS["high_row"],
    "Medium":   COLOURS["medium_row"],
    "Low":      COLOURS["low_row"],
}

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _make_header_fill(hex_colour: str) -> PatternFill:
    """Create a solid PatternFill for header cells."""
    return PatternFill(fill_type="solid", fgColor=hex_colour)


def _make_header_font(bold: bool = True) -> Font:
    """Create white, bold header font."""
    return Font(name="Calibri", bold=bold, color=COLOURS["header_font"], size=11)


def _make_title_font(size: int = 14) -> Font:
    """Create large bold title font in navy."""
    return Font(name="Calibri", bold=True, color=COLOURS["navy"], size=size)


def _auto_fit_columns(ws, min_width: int = 10, max_width: int = 60) -> None:
    """Iterate all columns and set width to the longest cell value.

    Args:
        ws: openpyxl Worksheet object.
        min_width: Minimum column width in character units.
        max_width: Maximum column width to prevent overly wide columns.
    """
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        adjusted = min(max(max_len + 4, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _write_kpi_tile(
    ws,
    row: int,
    col: int,
    label: str,
    value: str,
    colour: str,
) -> None:
    """Write a two-cell KPI tile (label above, value below) with styling.

    Args:
        ws: openpyxl Worksheet.
        row: Starting row for the label cell.
        col: Column for the tile.
        label: KPI name string.
        value: KPI value string.
        colour: Hex colour for the label background.
    """
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.fill = _make_header_fill(colour)
    label_cell.font = _make_header_font()
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    label_cell.border = THIN_BORDER

    value_cell = ws.cell(row=row + 1, column=col, value=value)
    value_cell.font = Font(name="Calibri", bold=True, size=16, color=colour)
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.fill = _make_header_fill(COLOURS["light_grey"])
    value_cell.border = THIN_BORDER


def _build_summary_sheet(
    ws,
    summary: dict[str, Any],
    run_date: str,
) -> None:
    """Populate the Summary sheet with KPI tiles and breakdown tables.

    Args:
        ws: openpyxl Worksheet (Summary tab).
        summary: Executive summary dict from scorer.build_executive_summary().
        run_date: ISO date string for report header.
    """
    ws.sheet_properties.tabColor = COLOURS["navy"]
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    # Title block
    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = "OPERATIONS COST LEAKAGE DETECTOR — EXECUTIVE SUMMARY"
    title.font = Font(name="Calibri", bold=True, size=16, color=COLOURS["white"])
    title.fill = _make_header_fill(COLOURS["navy"])
    title.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:H2")
    sub = ws["A2"]
    sub.value = (
        f"Report Date: {run_date}  |  "
        f"Analysis Period: {summary['date_range']['start']} → {summary['date_range']['end']}  |  "
        f"Organisation: Acme Operations Ltd"
    )
    sub.font = Font(name="Calibri", italic=True, size=10, color=COLOURS["navy"])
    sub.alignment = Alignment(horizontal="center", vertical="center")

    # --- KPI Tiles row 4–5 ---
    sev = summary["severity_breakdown"]
    kpi_tiles = [
        ("TOTAL LEAKAGE",    f"£{summary['headline_gbp']:,.2f}",       COLOURS["dark_red"]),
        ("TRANSACTIONS",     f"{summary['total_transactions_analysed']:,}",  COLOURS["navy"]),
        ("FLAGS RAISED",     f"{summary['total_flags']:,}",             COLOURS["navy"]),
        ("CRITICAL",         str(sev.get("Critical", 0)),               "CC0000"),
        ("HIGH",             str(sev.get("High", 0)),                   "C65911"),
        ("MEDIUM",           str(sev.get("Medium", 0)),                 COLOURS["gold"]),
        ("LOW",              str(sev.get("Low", 0)),                    COLOURS["dark_green"]),
    ]
    for i, (label, value, colour) in enumerate(kpi_tiles, start=1):
        _write_kpi_tile(ws, row=4, col=i, label=label, value=value, colour=colour)
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 30

    # --- Leakage by Rule table (row 7) ---
    ws.cell(row=7, column=1, value="LEAKAGE BY DETECTION RULE").font = _make_title_font(12)
    headers = ["Rule", "Flags", "Est. Leakage (£)", "% of Total"]
    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col_i, value=h)
        cell.fill = _make_header_fill(COLOURS["navy"])
        cell.font = _make_header_font()
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    rule_display = {
        "duplicate":      "Duplicate Transactions",
        "price_variance": "Price Variance / Overcharge",
        "sla_breach":     "SLA Breach",
        "volume_spike":   "Volume Spike",
    }
    total_leakage = summary["headline_gbp"] or 1  # avoid div-by-zero
    for row_i, (rule, data) in enumerate(summary["by_rule"].items(), start=9):
        leakage = data.get("leakage_gbp", 0)
        pct = (leakage / total_leakage * 100) if total_leakage else 0
        row_data = [
            rule_display.get(rule, rule),
            data.get("count", 0),
            round(leakage, 2),
            f"{pct:.1f}%",
        ]
        for col_i, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.fill = _make_header_fill(COLOURS["light_grey"])
            cell.border = THIN_BORDER
            if col_i in (2, 3):
                cell.number_format = "#,##0.00" if col_i == 3 else "#,##0"
            cell.alignment = Alignment(horizontal="right" if col_i > 1 else "left")

    # --- Top 5 Suppliers table (col 6) ---
    ws.cell(row=7, column=6, value="TOP 5 SUPPLIERS BY LEAKAGE").font = _make_title_font(12)
    sup_headers = ["Supplier", "Est. Leakage (£)"]
    for col_i, h in enumerate(sup_headers, start=6):
        cell = ws.cell(row=8, column=col_i, value=h)
        cell.fill = _make_header_fill(COLOURS["dark_red"])
        cell.font = _make_header_font()
        cell.border = THIN_BORDER

    for row_i, (sup, val) in enumerate(summary["top_suppliers"].items(), start=9):
        ws.cell(row=row_i, column=6, value=sup).border = THIN_BORDER
        amt_cell = ws.cell(row=row_i, column=7, value=round(val, 2))
        amt_cell.number_format = "#,##0.00"
        amt_cell.border = THIN_BORDER

    _auto_fit_columns(ws)


def _build_flagged_sheet(ws, scored: pd.DataFrame) -> None:
    """Write all flagged transactions with severity-coded row colours.

    Args:
        ws: openpyxl Worksheet (Flagged Items tab).
        scored: Scored DataFrame from scorer.score_flagged_transactions().
    """
    ws.sheet_properties.tabColor = COLOURS["dark_red"]

    display_cols = [
        "transaction_id", "date", "supplier_name", "category", "region",
        "invoice_amount", "baseline_rate", "leakage_amount_gbp",
        "rule_triggered", "severity", "composite_score",
        "rule_detail", "action_required", "approved_by",
    ]
    # Only include columns that exist
    display_cols = [c for c in display_cols if c in scored.columns]
    df_display = scored[display_cols].copy()

    # Friendly column headers
    col_rename = {
        "transaction_id":    "Transaction ID",
        "date":              "Date",
        "supplier_name":     "Supplier",
        "category":          "Category",
        "region":            "Region",
        "invoice_amount":    "Invoice Amt (£)",
        "baseline_rate":     "Baseline Rate (£)",
        "leakage_amount_gbp":"Leakage Est. (£)",
        "rule_triggered":    "Rule",
        "severity":          "Severity",
        "composite_score":   "Score",
        "rule_detail":       "Detail",
        "action_required":   "Action Required",
        "approved_by":       "Approved By",
    }
    df_display = df_display.rename(columns=col_rename)

    # Write header row
    header_row = list(df_display.columns)
    for col_i, h in enumerate(header_row, start=1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.fill = _make_header_fill(COLOURS["dark_red"])
        cell.font = _make_header_font()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # Write data rows with severity colouring
    severity_col_idx = header_row.index("Severity") + 1 if "Severity" in header_row else None

    for row_i, row in enumerate(
        dataframe_to_rows(df_display, index=False, header=False), start=2
    ):
        severity = str(row[severity_col_idx - 1]) if severity_col_idx else "Low"
        row_colour = SEVERITY_ROW_COLOURS.get(severity, COLOURS["light_grey"])
        fill = _make_header_fill(row_colour)

        for col_i, val in enumerate(row, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=False, vertical="center")
            # Format currency columns
            col_name = header_row[col_i - 1]
            if "(£)" in col_name:
                cell.number_format = "#,##0.00"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header_row))}1"
    _auto_fit_columns(ws)


def _build_statistics_sheet(ws, scored: pd.DataFrame) -> None:
    """Write rule performance metrics and a daily leakage trend table.

    Args:
        ws: openpyxl Worksheet (Statistics tab).
        scored: Scored DataFrame.
    """
    ws.sheet_properties.tabColor = COLOURS["dark_green"]

    # --- Rule metrics ---
    ws.cell(row=1, column=1, value="RULE PERFORMANCE METRICS").font = _make_title_font()

    rule_headers = [
        "Rule", "Total Flags", "Critical", "High", "Medium", "Low",
        "Total Leakage (£)", "Avg Leakage / Flag (£)", "Max Leakage (£)",
    ]
    for col_i, h in enumerate(rule_headers, start=1):
        cell = ws.cell(row=2, column=col_i, value=h)
        cell.fill = _make_header_fill(COLOURS["dark_green"])
        cell.font = _make_header_font()
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    rule_display = {
        "duplicate":      "Duplicate Transactions",
        "price_variance": "Price Variance",
        "sla_breach":     "SLA Breach",
        "volume_spike":   "Volume Spike",
    }

    for row_i, (rule, grp) in enumerate(
        scored.groupby("rule_triggered"), start=3
    ):
        sev_counts = grp["severity"].value_counts()
        total_leak = grp["leakage_amount_gbp"].sum()
        row_data = [
            rule_display.get(rule, rule),
            len(grp),
            sev_counts.get("Critical", 0),
            sev_counts.get("High", 0),
            sev_counts.get("Medium", 0),
            sev_counts.get("Low", 0),
            round(total_leak, 2),
            round(total_leak / len(grp), 2) if len(grp) > 0 else 0,
            round(grp["leakage_amount_gbp"].max(), 2),
        ]
        for col_i, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.fill = _make_header_fill(COLOURS["light_grey"])
            cell.border = THIN_BORDER
            if col_i > 6:
                cell.number_format = "#,##0.00"

    # --- Daily leakage trend table ---
    scored_with_date = scored.copy()
    if not pd.api.types.is_datetime64_any_dtype(scored_with_date["date"]):
        scored_with_date["date"] = pd.to_datetime(scored_with_date["date"])

    daily = (
        scored_with_date.groupby(scored_with_date["date"].dt.date)
        .agg(
            flags=("transaction_id", "count"),
            leakage_gbp=("leakage_amount_gbp", "sum"),
            critical=("severity", lambda s: (s == "Critical").sum()),
        )
        .reset_index()
        .rename(columns={"date": "Date"})
    )

    start_row = 10
    ws.cell(row=start_row, column=1, value="DAILY LEAKAGE TREND").font = _make_title_font()
    daily_headers = ["Date", "Flags", "Critical", "Total Leakage (£)"]
    for col_i, h in enumerate(daily_headers, start=1):
        cell = ws.cell(row=start_row + 1, column=col_i, value=h)
        cell.fill = _make_header_fill(COLOURS["navy"])
        cell.font = _make_header_font()
        cell.border = THIN_BORDER

    for row_i, row in enumerate(
        dataframe_to_rows(daily, index=False, header=False),
        start=start_row + 2,
    ):
        for col_i, val in enumerate(row, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.fill = _make_header_fill(COLOURS["light_grey"])
            cell.border = THIN_BORDER
            if col_i == 4:
                cell.number_format = "#,##0.00"

    # Embed a simple bar chart: daily leakage
    if len(daily) > 1:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Daily Estimated Leakage (£)"
        chart.y_axis.title = "£ Leakage"
        chart.x_axis.title = "Date"
        chart.style = 10
        chart.height = 12
        chart.width = 22

        data_ref = Reference(
            ws,
            min_col=4,
            min_row=start_row + 1,
            max_row=start_row + 1 + len(daily),
        )
        chart.add_data(data_ref, titles_from_data=True)
        ws.add_chart(chart, "F10")

    _auto_fit_columns(ws)


def generate_report(
    scored: pd.DataFrame,
    summary: dict[str, Any],
    config_path: str = "config.yaml",
) -> Path:
    """Generate the full Excel workbook and write it to the output directory.

    Args:
        scored: Scored and classified flagged transactions.
        summary: Executive summary dict from scorer.build_executive_summary().
        config_path: Path to configuration YAML.

    Returns:
        Path to the generated .xlsx file.

    Raises:
        OSError: If output directory cannot be created.
    """
    with open(config_path, "r") as fh:
        cfg = yaml.safe_load(fh)

    run_date = datetime.today().strftime("%Y-%m-%d")
    output_dir = Path(cfg["paths"]["output_dir"])
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = cfg["paths"]["report_filename"].format(date=run_date)
    output_path = output_dir / filename

    wb = Workbook()
    wb.remove(wb.active)  # Remove default blank sheet

    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Summary")
    _build_summary_sheet(ws_summary, summary, run_date)
    logger.info("Built Summary sheet")

    # Sheet 2: Flagged Items
    ws_flagged = wb.create_sheet("Flagged Items")
    _build_flagged_sheet(ws_flagged, scored)
    logger.info("Built Flagged Items sheet (%d rows)", len(scored))

    # Sheet 3: Statistics
    ws_stats = wb.create_sheet("Statistics")
    _build_statistics_sheet(ws_stats, scored)
    logger.info("Built Statistics sheet")

    wb.save(output_path)
    logger.info("Excel report saved to %s", output_path)
    return output_path
